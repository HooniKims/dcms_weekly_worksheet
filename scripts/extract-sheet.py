#!/usr/bin/env python3
import argparse
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def serializable(value: object) -> object:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return ""
    return value


parser = argparse.ArgumentParser()
parser.add_argument("input", type=Path)
parser.add_argument("output", type=Path)
args = parser.parse_args()

workbook = load_workbook(args.input, read_only=True, data_only=False)
if "제출원본" not in workbook.sheetnames:
    raise RuntimeError("missing-sheet:제출원본")

raw_sheet = workbook["제출원본"]
values = [
    [serializable(cell) for cell in row]
    for row in raw_sheet.iter_rows(min_row=1, max_col=7, values_only=True)
]
while values and all(value == "" for value in values[-1]):
    values.pop()

payload = {
    "sheetTitles": workbook.sheetnames,
    "values": values,
}
args.output.parent.mkdir(parents=True, exist_ok=True)
args.output.write_text(
    json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
    encoding="utf-8",
)
print(
    json.dumps(
        {
            "sheets": len(workbook.sheetnames),
            "rows": len(values),
            "output": str(args.output),
        },
        ensure_ascii=False,
    )
)
